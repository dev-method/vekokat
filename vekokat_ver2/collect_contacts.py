# -*- coding: utf-8 -*-
import os, sys, re
sys.path.insert(0, '/usr/src/vekokat.ru/vekokat_ver2')
sys.path.append('/usr/src/vekokat.ru/vekokat_ver2')
os.environ.setdefault("DJANGO_SETTINGS_MODULE", 'vekokat_ver2.settings')
import django
django.setup()
from django.core.mail import send_mail
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.drawing.image import Image

wb = load_workbook('/usr/src/vekokat.ru/vekokat_ver2/static/core/images/contact_list_starter.xlsx')
ws = wb.active
try:
  os.remove('/usr/src/vekokat.ru/vekokat_ver2/static/core/images/contact_list.xlsx')
except:
  pass

from orders.models import Order, ServiceOrder, NewOrder
messages=ServiceOrder.objects.all()
feedbacks=NewOrder.objects.all()
srvord_contacts = {}
nword_contacts = {}

for item in messages:
    if item.name:
        name=item.name.strip()
    else:
        name="не указано"
    item_strip=item.contacts.strip()
    if re.match(r'^((8|\+7)[\- ]?)?(\(?\d{3}\)?[\- ]?)?[\d\- ]{7,10}$', item_strip):
        srvord_contacts[item.id]={"name":name, "phone":item_strip, "mail":"не указана"}
    elif re.match(r'(^|\s)[-a-z0-9_.]+@([-a-z0-9]+\.)+[a-z]{2,6}(\s|$)', item_strip.lower()):
        srvord_contacts[item.id] = {"name": name, "phone": "не указана", "mail": item_strip }

for item in feedbacks:
    phone=""
    mail=""
    phone_flag=True
    mail_flag=True
    if item.name:
        name=item.name.strip()
    else:
        name="не указано"
    if item.phone:
        item_phone_strip=item.phone.strip()
        if re.match(r'^((8|\+7)[\- ]?)?(\(?\d{3}\)?[\- ]?)?[\d\- ]{7,10}$', item_phone_strip):
          phone= item_phone_strip
          phone_flag=True
    else:
        phone = "не указан"
        phone_flag = False
    if item.mail:
        item_mail_strip=item.mail.strip()
        if re.match(r'(^|\s)[-a-z0-9_.]+@([-a-z0-9]+\.)+[a-z]{2,6}(\s|$)', item_mail_strip.lower()):
          mail= item_mail_strip
          mail_flag = True
    else:
        mail = "не указана"
        mail_flag=False
    if phone_flag or mail_flag:
        nword_contacts[item.id] = {"name": name, "phone": phone, "mail": mail}

counter=4
num=1
srvord_keys=srvord_contacts.keys()
for item in srvord_keys:
    item_dict=srvord_contacts[item]
    refNum = ws['A' + str(counter)]
    refNum.value = str(num)
    refName=ws['B'+str(counter)]
    refName.value = item_dict["name"]
    refPhone = ws['C' + str(counter)]
    refPhone.value = item_dict["phone"]
    refMail = ws['D' + str(counter)]
    refMail.value = item_dict["mail"]
    counter+=1
    num+=1

nword_keys=nword_contacts.keys()
for item in nword_keys:
    item_dict=nword_contacts[item]
    refNum = ws['A' + str(counter)]
    refNum.value = str(num)
    refName=ws['B'+str(counter)]
    refName.value = item_dict["name"]
    refPhone = ws['C' + str(counter)]
    refPhone.value = item_dict["phone"]
    refMail = ws['D' + str(counter)]
    refMail.value = item_dict["mail"]
    counter+=1
    num+=1

wb.save('/usr/src/vekokat.ru/vekokat_ver2/static/core/images/contact_list.xlsx')
mail_subject="Обновленный список контактов всех, кто обращался через формы на сайте: www.vekokat.ru/static/core/images/contact_list.xlsx"
send_mail('Заявка', mail_subject, 'admin@vekomet.ru',['vekokat@gmail.com'], fail_silently=False)

print ('excellent')
